import os
import json
import csv
import re
import pandas as pd
from pathlib import Path
import logging
import shutil
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional, Tuple
from app.core.config import get_settings
from app.services.article_parser import extract_article_code, get_hierarchy_level

settings = get_settings()
logger = logging.getLogger(__name__)

# Keywords used to identify the primary article/code column
ARTICLE_KEYWORDS = ["article", "code", "réf", "n°", "art.", "item", "n. index", "articles", "related names"]
TECH_KEYWORDS = ["qt", "p.u", "somme", "prix", "unit", "montant", "total", "qte", "quantité", "ht", "tva", "u", "pu", "unite"]

def normalize_number(x: Any) -> Optional[float]:
    """Normalize numeric strings by handling French currency and decimal separators."""
    if x is None or x == "" or (isinstance(x, str) and x.lower() == "nan"):
        return None
    
    # Clean string: remove currency symbols and f8fafcspace
    s = str(x).strip()
    s = re.sub(r'[€$£%]', '', s)
    s = re.sub(r'\s+', '', s)
    
    # Handle thousands and decimal separators
    if ',' in s and '.' in s:
        # Determine if French (1.200,50) or English (1,200.50)
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        # Common French decimal or English thousands
        s = s.replace(',', '.')
    elif s.count('.') > 1:
        # Multiple dots indicate thousands separators (e.g. 1.000.000)
        s = s.replace('.', '')
        
    try:
        val = float(s)
        if val.is_integer():
            return int(val)
        return val
    except (ValueError, TypeError):
        return None

def is_val_missing(val: Any) -> bool:
    """Comprehensive check for empty, zero, or placeholder values."""
    s = str(val).strip().lower()
    return not s or s in ["nan", "none", "0", "0.0", "0,00", "0.00", "0,00 €", "-", "/", "null"]

def list_session_files(session_id: str) -> List[str]:
    """
    Lists all files available in the upload directory for a specific session.
    """
    session_dir = Path(settings.upload_dir) / session_id
    if not session_dir.exists() or not session_dir.is_dir():
        return []
    
    return [f.name for f in session_dir.iterdir() if f.is_file()]

def is_header_row(values: List[Any]) -> bool:
    """Universal check for a header row based on keywords."""
    vals = [str(v).lower() for v in values if v is not None]
    has_art = any(k in vals for k in ARTICLE_KEYWORDS) or any(k in " ".join(vals) for k in ["n. index", "art."])
    has_tech = any(k in " ".join(vals) for k in TECH_KEYWORDS)
    return has_art and has_tech

def manage_csv_data(session_id: str, filename: str, action: str = "read", updates: Optional[List[Dict[str, Any]]] = None) -> str:
    """
    Manages CSV data.
    - 'read': returns structure and missing field info.
    - 'update': fills missing fields in existing rows (NEVER overwrites existing data).
    - 'insert': adds new rows at a specific index to maintain serial order.
    """
    base_dir = Path(settings.upload_dir) / session_id
    is_excel = filename.lower().endswith('.xlsx')
    
    # Versioning Logic: Always prefer the most recently updated version as the source
    current_filename = filename
    if not filename.startswith("updated_"):
        potential_updated = base_dir / f"updated_{filename}"
        if potential_updated.exists():
            current_filename = f"updated_{filename}"
            logger.info(f"CSV_VERSIONING | Using latest version: {current_filename}")

    file_path = base_dir / current_filename
    
    if not file_path.exists():
        return f"Error: File '{filename}' not found."

    # Ensure the directory for updates exists
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        if is_excel:
            # Detect header row for Excel by scanning first 20 rows
            header_check = pd.read_excel(file_path, nrows=20, header=None, engine='openpyxl')
            header_idx = 0
            for i, row in header_check.iterrows():
                if is_header_row(row.values):
                    header_idx = i
                    break
            
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl', header=header_idx)
            encoding = 'utf-8' # Not used for writing excel but for consistency
        else:
            # Robust encoding and separator detection for CSV
            encoding = None
            sep = ','
            for enc in ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        sample = f.read(8192)
                        if not sample: continue
                        dialect = csv.Sniffer().sniff(sample)
                        sep = dialect.delimiter
                        encoding = enc
                        break
                except Exception:
                    continue
            
            if not encoding:
                encoding = 'utf-8'
                sep = ';' if ';' in open(file_path, 'r', encoding='utf-8', errors='replace').read(1024) else ','

            # Header detection: identify metadata rows to preserve design
            header_idx = 0
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                for i, line in enumerate(f):
                    if is_header_row(line.split(sep)):
                        header_idx = i
                        break
            
            # Load CSV starting from the detected header row
            df = pd.read_csv(file_path, encoding=encoding, sep=sep, engine='python', dtype=str, header=header_idx)

        if action == "insert" and updates:
            # 'updates' contains objects with 'row' (insertion index) and 'data' (dict)
            new_filename = filename if filename.startswith("updated_") else f"updated_{filename}"
            new_file_path = file_path.parent / new_filename
            
            if is_excel:
                if file_path != new_file_path:
                    shutil.copy2(file_path, new_file_path)
                wb = load_workbook(new_file_path)
                ws = wb.active
                # Map column names to indices
                col_map = {col: i + 1 for i, col in enumerate(df.columns)}
                
                # Sort updates by row descending to prevent index shift during insertion
                for upd in sorted(updates, key=lambda x: int(x['row']), reverse=True):
                    idx = int(upd['row']) + header_idx + 2
                    ws.insert_rows(idx)
                    for col_name, val in upd.get('data', {}).items():
                        if col_name in col_map:
                            ws.cell(row=idx, column=col_map[col_name]).value = val
                wb.save(new_file_path)
                
                return json.dumps({"status": "success", "message": f"Successfully inserted {len(updates)} rows into {new_filename}."})

            else:
                # CSV Insertion logic
                for upd in sorted(updates, key=lambda x: int(x['row']), reverse=True):
                    idx = int(upd['row'])
                    new_row = pd.Series(upd.get('data', {}))
                    df = pd.concat([df.iloc[:idx], new_row.to_frame().T, df.iloc[idx:]]).reset_index(drop=True)
                
                with open(new_file_path, 'w', encoding=encoding, newline='') as f:
                    source_for_headers = file_path
                    with open(source_for_headers, 'r', encoding=encoding, errors='replace') as old_f:
                        for _ in range(header_idx):
                            f.write(old_f.readline())
                    
                    df.to_csv(f, index=False, encoding=encoding, sep=sep, quoting=csv.QUOTE_MINIMAL, na_rep="")
                return json.dumps({"status": "success", "message": f"Inserted {len(updates)} rows into CSV {filename}"})

        if action == "update" and updates:
            # Apply updates provided by the LLM
            new_filename = filename if filename.startswith("updated_") else f"updated_{filename}"
            new_file_path = file_path.parent / new_filename
            article_col = next((c for c in df.columns if any(k in str(c).lower() for k in ARTICLE_KEYWORDS)), None)
            
            logger.info(f"CSV_UPDATE_EXECUTION | File: {filename} | Updates: {len(updates)}")
            
            # Excel specific logic to preserve formatting
            if is_excel:
                # 1. Copy original file to the new path to preserve formatting/images/styles
                if file_path != new_file_path:
                    shutil.copy2(file_path, new_file_path)
                wb = load_workbook(new_file_path)
                ws = wb.active
                # Map column names to 1-based indices
                col_map = {col: i + 1 for i, col in enumerate(df.columns)}
                
            for upd in updates:
                try:
                    row_idx = int(upd.get("row"))
                except (ValueError, TypeError):
                    continue

                col_name = upd.get("column")
                raw_val = upd.get("value")
                
                norm_val = normalize_number(raw_val)
                
                if norm_val is not None:
                    val_to_assign = norm_val
                elif raw_val is not None and str(raw_val).strip().lower() not in ["noos", "none", "nan", "null", ""]:
                    val_to_assign = str(raw_val)
                else:
                    val_to_assign = None 
                
                if col_name not in df.columns:
                    logger.warning(f"SKIP_NEW_COL | {col_name} is not in original CSV. Skipping.")
                    continue

                if row_idx < len(df):
                    # NO OVERWRITE PROTECTION: Check if current cell is already filled
                    existing_val = df.at[row_idx, col_name]
                    if not is_val_missing(existing_val):
                        logger.info(f"PROTECTION | Row {row_idx}, Col {col_name} already has value. Skipping.")
                        continue

                    # Safety: Skip metadata rows or chapter titles
                    if article_col:
                        row_val = str(df.at[row_idx, article_col]).strip()
                        # Skip only high-level headers. Allow items even if they contain 'total' unless it's the absolute final row.
                        is_doc_metadata = any(k in row_val.lower() for k in ["chantier", "client", "projet", "devis n"])
                        is_final_total = "total général" in row_val.lower() or row_val.lower() == "total"

                        if not row_val or row_val.lower() in ["nan", "none"] or is_doc_metadata or is_final_total:
                            continue

                    if is_excel:
                        # Excel row = df_idx + header_idx + 1 (1-indexed) + 1 (data starts after header)
                        excel_row = row_idx + header_idx + 2
                        excel_col = col_map.get(col_name)
                        if excel_col:
                            ws.cell(row=excel_row, column=excel_col).value = val_to_assign
                            logger.info(f"EXCEL_CELL_FILL: {col_name} at Row {excel_row}")
                    else:
                        # Fallback for CSV
                        df.at[row_idx, col_name] = val_to_assign

            if is_excel:
                wb.save(new_file_path)
                return json.dumps({"status": "success", "message": f"Successfully updated {len(updates)} fields in {new_filename}."})
            else:
                # Standard CSV save logic
                with open(new_file_path, 'w', encoding=encoding, newline='') as f:
                    # If updating the same file, we need to read from the original
                    source_for_headers = file_path
                    if filename.startswith("updated_") and (base_dir / filename.replace("updated_", "")).exists():
                         source_for_headers = base_dir / filename.replace("updated_", "")
                    with open(source_for_headers, 'r', encoding=encoding, errors='replace') as old_f:
                        for _ in range(header_idx):
                            f.write(old_f.readline())
                    df.to_csv(f, index=False, encoding=encoding, sep=sep, quoting=csv.QUOTE_MINIMAL, na_rep="", float_format='%.10g')
                return json.dumps({"status": "success", "message": f"Successfully updated {len(updates)} fields in CSV {new_filename}."})
            
        # Global Missing Field Detection
        # Identifies terminal articles (3+ segments) with empty technical columns
        missing_audit = []
        article_col = next((c for c in df.columns if any(k in str(c).lower() for k in ARTICLE_KEYWORDS)), None)
        tech_cols = [c for c in df.columns if any(k in str(c).lower() for k in TECH_KEYWORDS)]
        
        if article_col:
            for idx, row in df.iterrows():
                art_val = str(row[article_col]).strip()
                # Detect if the row is a valid article (at least one numeric segment)
                code = extract_article_code(art_val)
                is_doc_metadata = any(k in art_val.lower() for k in ["chantier", "client", "projet", "devis n"])
                is_final_total = "total général" in art_val.lower() or art_val.lower() == "total"
                
                if code and not is_doc_metadata and not is_final_total:
                    missing_in_row = [c for c in tech_cols if is_val_missing(row[c])]
                    if missing_in_row:
                        missing_audit.append({
                            "row": idx, 
                            "article": art_val, 
                            "code": code,
                            "level": get_hierarchy_level(code),
                            "missing_columns": missing_in_row
                        })

        # Return all columns so LLM can see 'Unnamed' columns that might contain data
        valid_columns = list(df.columns)
        sample_data = df.head(5).to_dict(orient="records") # Reduced from 50 to 5 to save tokens

        summary = {
            "status": "success",
            "current_file": current_filename,
            "columns": valid_columns,
            "row_count": len(df),
            "missing_fields_to_fix": missing_audit, # Full list of all missing fields in the entire file
            "sample_data": sample_data,
            "instruction": (
                "CRITICAL: You must iterate through EVERY single item in 'missing_fields_to_fix'. "
                "Do not skip any articles. Process them in order. "
                "2. Use the 'row' index provided. "
                "3. Cross-reference with 'list_all_pdf_articles' to find PDF articles not in the CSV."
            )
        }
        return json.dumps(summary, ensure_ascii=False)
    except ImportError as ie:
        error_msg = f"Dependency error: 'openpyxl' is required for Excel files. Please install it."
        logger.error(f"TOOL_DEPS_ERROR | {filename}: {str(ie)}")
        return json.dumps({"status": "error", "message": error_msg})
    except Exception as e:
        logger.error(f"TOOL_ERROR | {filename}: {str(e)}")
        return json.dumps({"status": "error", "message": f"Technical error reading {filename}: {str(e)}. Ensure the file is not corrupted and uses standard CSV/Excel formatting."})
